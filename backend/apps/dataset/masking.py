"""列级脱敏（v0.27）：对指定列的值做遮蔽。管理员/属主看原值，普通用户看脱敏值。"""

from __future__ import annotations


def _mask_value(value, strategy: str) -> str:
    if value is None:
        return ""
    s = str(value)
    if strategy == "full" or len(s) <= 2:
        return "*" * max(len(s), 4)
    if strategy == "partial":
        return s[0] + "*" * (len(s) - 2) + s[-1]
    return "*" * len(s)


def mask_rows(columns: list[str], rows, rules: dict[str, str]):
    """对 rows 中、列名命中 rules 的单元格脱敏。

    :param rules: {列名: 策略}。返回脱敏后的新行列表。
    """
    if not rules:
        return list(rows)
    idx_strategy = {
        i: rules[col] for i, col in enumerate(columns) if col in rules
    }
    if not idx_strategy:
        return list(rows)
    out = []
    for row in rows:
        new = list(row)
        for i, strategy in idx_strategy.items():
            if i < len(new):
                new[i] = _mask_value(new[i], strategy)
        out.append(new)
    return out


def rules_for(dataset) -> dict[str, str]:
    """取数据集的脱敏规则 {列名: 策略}。"""
    from apps.dataset.models import MaskingRule

    return {
        r.column: r.strategy
        for r in MaskingRule.objects.filter(dataset=dataset)
    }


def stream_filtered_xlsx(src_path, *, columns=None, rules=None):
    """流式读源 xlsx →（可选选列 + 脱敏）→ 写入内存 BytesIO（B1/B3 修复）。

    用 openpyxl read_only + write_only 逐行处理，**不落临时文件、内存占用恒定**。
    返回可直接交给 FileResponse 的 BytesIO。
    """
    from io import BytesIO

    from openpyxl import Workbook, load_workbook

    rules = rules or {}
    wb_in = load_workbook(str(src_path), read_only=True, data_only=True)
    try:
        ws_in = wb_in.active
        rows_iter = ws_in.iter_rows(values_only=True)
        try:
            raw_header = list(next(rows_iter))
        except StopIteration:
            raw_header = []
        header = ["" if h is None else str(h) for h in raw_header]
        idx = [header.index(c) for c in columns if c in header] if columns else list(
            range(len(header))
        )
        out_header = [header[i] for i in idx]
        mask_pos = {
            pos: rules[col] for pos, col in enumerate(out_header) if col in rules
        }

        wb_out = Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title="数据")
        ws_out.append(out_header)
        for row in rows_iter:
            vals = [row[i] if i < len(row) else None for i in idx]
            for pos, strategy in mask_pos.items():
                vals[pos] = _mask_value(vals[pos], strategy)
            ws_out.append(vals)

        buf = BytesIO()
        wb_out.save(buf)
        buf.seek(0)
        return buf
    finally:
        wb_in.close()
