"""Excel 导出：openpyxl 流式写入，支持大结果集。

使用 write_only 模式逐行写，避免一次性把全部数据放进内存。
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

# 以这些字符开头的文本，Excel/WPS 打开时会被当公式执行（CSV/Excel 注入）。
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize(value: object) -> object:
    """对疑似公式的字符串单元格加前导单引号，使其被当作纯文本（防注入）。

    只处理字符串；数字/日期/None 等原样返回，不影响数据类型与展示。
    """
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def read_xlsx(
    path: str | Path, *, limit: int | None = None, columns: list[str] | None = None
) -> tuple[list[str], list[list[object]]]:
    """流式读回 xlsx（read_only）：返回 (列名, 行)。

    :param limit: 最多读取的数据行数（None=全部）。用于在线预览。
    :param columns: 只取这些列（按表头名，保持给定顺序）；None=全部列。用于选列下载。
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return [], []
        header = ["" if h is None else str(h) for h in header]
        if columns:
            idx = [header.index(c) for c in columns if c in header]
            out_header = [header[i] for i in idx]
        else:
            idx = list(range(len(header)))
            out_header = header
        out_rows: list[list[object]] = []
        for n, row in enumerate(rows_iter):
            if limit is not None and n >= limit:
                break
            out_rows.append([row[i] if i < len(row) else None for i in idx])
        return out_header, out_rows
    finally:
        wb.close()


def export_to_xlsx(
    columns: Sequence[str],
    rows: Iterable[Sequence[object]],
    output_path: str | Path,
    sheet_name: str = "数据",
) -> int:
    """把表头 + 数据行写入 xlsx，返回写入的数据行数。

    :param columns: 列名列表（表头）。
    :param rows: 数据行的可迭代对象（可为生成器，实现流式）。
    :param output_path: 输出文件路径。
    :param sheet_name: 工作表名。
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet_name)

    # 表头加粗
    header_font = Font(bold=True)
    header_cells = []
    for name in columns:
        cell = WriteOnlyCell(worksheet, value=name)
        cell.font = header_font
        header_cells.append(cell)
    worksheet.append(header_cells)

    row_count = 0
    for row in rows:
        worksheet.append([_sanitize(v) for v in row])
        row_count += 1

    workbook.save(str(output_path))
    return row_count
