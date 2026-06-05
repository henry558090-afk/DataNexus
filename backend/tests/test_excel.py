"""excel 导出单元测试。"""

import openpyxl

from core.excel import export_to_xlsx


def test_export_writes_header_and_rows(tmp_path):
    path = tmp_path / "out.xlsx"
    count = export_to_xlsx(["name", "amount"], [["a", 1], ["b", 2]], path)

    assert count == 2

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    assert worksheet["A1"].value == "name"
    assert worksheet["B1"].value == "amount"
    assert worksheet["A2"].value == "a"
    assert worksheet["B3"].value == 2


def test_export_empty_rows_writes_header_only(tmp_path):
    # 0 行结果也要有表头（修复 B 回归）
    path = tmp_path / "empty.xlsx"
    count = export_to_xlsx(["a", "b"], [], path)
    assert count == 0
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    assert worksheet["A1"].value == "a"
    assert worksheet["B1"].value == "b"
    assert worksheet.max_row == 1  # 仅表头


def test_export_accepts_generator(tmp_path):
    path = tmp_path / "gen.xlsx"

    def rows():
        for i in range(5):
            yield [i, i * i]

    count = export_to_xlsx(["n", "n2"], rows(), path)
    assert count == 5
