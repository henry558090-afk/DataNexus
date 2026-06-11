"""v0.28 缺陷修复回归：B1/B3 流式下载、B2 有界池、B3/B6 图表、B4 参数校验、B5 SSO 短码。"""

from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts import views as account_views
from apps.accounts.models import WecomLoginCode
from apps.catalog.models import Folder, FolderShare
from apps.dataset.models import Dataset, MaskingRule
from apps.datasource.models import DataSource
from apps.execution.models import DataFile
from core import excel

User = get_user_model()


@pytest.fixture
def manager(db):
    return User.objects.create_user("mgr", password="x", is_assistant_admin=True)


@pytest.fixture
def datasource(db):
    ds = DataSource(name="o", host="h", port=1521, service_name="s", username="u")
    ds.password = "p"
    ds.save()
    return ds


def cli(u) -> APIClient:
    c = APIClient()
    c.force_authenticate(user=u)
    return c


def _content(resp) -> bytes:
    return b"".join(resp.streaming_content) if resp.streaming else resp.content


# ---------- B1/B3 流式选列+脱敏下载（无临时文件） ----------
def test_masked_download_streams_masked_file(datasource, tmp_path, db):
    user = User.objects.create_user("v", password="x")
    folder = Folder.objects.create(name="客户")
    FolderShare.objects.create(folder=folder, subject_user=user)
    ds = Dataset.objects.create(
        name="客户", datasource=datasource, sql_text="SELECT 1", target_folder=folder
    )
    MaskingRule.objects.create(dataset=ds, column="手机", strategy="full")
    path = tmp_path / "c.xlsx"
    excel.export_to_xlsx(["姓名", "手机"], [["李四", "13900002222"]], path)
    f = DataFile.objects.create(
        dataset=ds,
        folder=folder,
        name="c.xlsx",
        status=DataFile.Status.SUCCESS,
        file_path=str(path),
    )
    resp = cli(user).get(f"/api/portal/files/{f.id}/download/")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(_content(resp)))
    ws = wb.active
    assert ws["A2"].value == "李四"
    assert set(ws["B2"].value) == {"*"}  # 手机被遮蔽


def test_column_select_download_streams_subset(datasource, tmp_path, manager):
    folder = Folder.objects.create(name="f")
    ds = Dataset.objects.create(
        name="d", datasource=datasource, sql_text="SELECT 1", target_folder=folder
    )
    path = tmp_path / "c.xlsx"
    excel.export_to_xlsx(["a", "b", "c"], [[1, 2, 3]], path)
    FolderShare.objects.create(folder=folder, subject_user=manager)
    f = DataFile.objects.create(
        dataset=ds,
        folder=folder,
        name="c.xlsx",
        status=DataFile.Status.SUCCESS,
        file_path=str(path),
    )
    resp = cli(manager).get(f"/api/portal/files/{f.id}/download/?columns=a,c")
    wb = load_workbook(BytesIO(_content(resp)))
    headers = [c.value for c in wb.active[1]]
    assert headers == ["a", "c"]


# ---------- B2 有界线程池 ----------
def test_run_executor_is_bounded(settings):
    settings.DATASET_RUN_CONCURRENCY = 3
    from apps.dataset import services

    services._run_pool = None  # 重置以应用新配置
    pool = services._run_executor()
    assert pool._max_workers == 3
    services._run_pool = None


# ---------- B3/B6 图表跳过非数值 ----------
def test_chart_skips_non_numeric(manager, datasource, tmp_path):
    folder = Folder.objects.create(name="f")
    ds = Dataset.objects.create(
        name="d", datasource=datasource, sql_text="SELECT 1", target_folder=folder
    )
    path = tmp_path / "c.xlsx"
    excel.export_to_xlsx(["部门", "金额"], [["销售", 10], ["销售", "abc"], ["财务", 5]], path)
    DataFile.objects.create(
        dataset=ds,
        folder=folder,
        name="c.xlsx",
        status=DataFile.Status.SUCCESS,
        file_path=str(path),
    )
    r = cli(manager).get(f"/api/datasets/{ds.id}/chart-data/?x=部门&y=金额&agg=sum")
    assert r.data["skipped"] == 1  # "abc" 被跳过而非当 0
    data = dict(zip(r.data["labels"], r.data["values"], strict=False))
    assert data["销售"] == 10  # 只算了有效的 10


# ---------- B4 参数占位符校验 ----------
def test_save_rejects_undeclared_param(manager, datasource):
    r = cli(manager).post(
        "/api/datasets/",
        {
            "name": "d",
            "datasource": datasource.id,
            "sql_text": "SELECT * FROM t WHERE dt = :dt",
        },
        format="json",
    )
    assert r.status_code == 400 and "dt" in str(r.data)


def test_save_accepts_declared_param(manager, datasource):
    r = cli(manager).post(
        "/api/datasets/",
        {
            "name": "d",
            "datasource": datasource.id,
            "sql_text": "SELECT * FROM t WHERE dt = :dt",
            "params": [{"name": "dt", "default": "2026-01-01"}],
        },
        format="json",
    )
    assert r.status_code == 201


def test_param_in_string_literal_not_flagged(manager, datasource):
    # 字符串里的 :x 不应被当成绑定变量
    r = cli(manager).post(
        "/api/datasets/",
        {"name": "d", "datasource": datasource.id, "sql_text": "SELECT 'a:b' FROM dual"},
        format="json",
    )
    assert r.status_code == 201


# ---------- B5 企微 SSO 一次性短码 ----------
def test_wecom_callback_uses_onetime_code(db, settings, monkeypatch):
    settings.WECOM_ENABLED = True
    settings.WECOM_CORP_ID = "corp"
    settings.WECOM_SECRET = "secret"
    settings.WECOM_REDIRECT_FRONTEND = "/"
    monkeypatch.setattr(account_views, "User", User)
    from apps.accounts import wecom

    monkeypatch.setattr(wecom, "exchange_code_for_userid", lambda code: "zhangsan")
    r = APIClient().get("/api/auth/wecom/callback/?code=abc")
    assert r.status_code == 302
    loc = r["Location"]
    assert "wecom_code=" in loc and "token=" not in loc  # token 不进 URL
    code = loc.split("wecom_code=")[1]
    # 用短码换 token
    ex = APIClient().post("/api/auth/wecom/exchange/", {"code": code}, format="json")
    assert ex.status_code == 200 and ex.data["token"]
    # 一次性：再换失败
    again = APIClient().post("/api/auth/wecom/exchange/", {"code": code}, format="json")
    assert again.status_code == 400
    assert not WecomLoginCode.objects.filter(code=code).exists()
